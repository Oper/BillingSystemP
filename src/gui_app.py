import calendar
import tkinter
import pandas as pd
from pathlib import Path
from openpyxl.reader.excel import load_workbook

from tkcalendar import DateEntry
from datetime import date, time, datetime
from tkinter import ttk, messagebox, filedialog
from tkinter.constants import END

from pydantic import ValidationError

from src.db.models import StatusClientEnum
from src.db.crud import delete_tariff, delete_client, get_client_by_pa, update_client, create_payment, create_client, \
    search_clients, get_clients, get_tariffs, create_tariff, get_tariff_by_name, set_client_activity, \
    get_payments_by_client, apply_monthly_charge, apply_daily_charge, get_accruals_by_client, create_accrual_daily, \
    create_accrual_monthly, get_debtors_report, set_client_status, get_last_payment_by_client, clear_db_clients, \
    bulk_create_clients, get_last_accrual_by_client, get_payments_in_range, get_payment_by_id, get_client_by_id, \
    create_service, get_services, get_service_by_name, delete_service
from src.db.database import get_db, init_db
from src.models.clients import ClientUpdate, ClientForPayments, ClientCard, ClientCreate, ClientBase
from src.models.payments import PaymentCreate
from src.models.tariffs import TariffCreate
from src.models.services import ServiceCreate


class BillingSysemApp(tkinter.Tk):
    """Основной класс приложения с графическим интерфейсом."""

    def __init__(self):
        super().__init__()
        self.title("Учет Клиентов Кабельного ТВ")
        self.geometry("800x600")
        self.resizable(width=False, height=False)

        # Стиль ttk (чтобы виджеты выглядели лучше)
        style = ttk.Style(self)
        # Попробуйте "clam", "alt", "default", "vista", "xpnative" (зависит от ОС)
        try:
            style.theme_use("vista")
        except tkinter.TclError:
            print("Тема 'vista' не найдена, используется 'default'")
        # Инициализация БД
        init_db()

        # Создание вкладок (Notebook)
        notebook = ttk.Notebook(self)
        notebook.pack(pady=10, padx=10, expand=True, fill="both")

        # Создание фреймов для вкладок
        abonents = ttk.Frame(notebook)
        tariffs = ttk.Frame(notebook)
        reports = ttk.Frame(notebook)
        settings = ttk.Frame(notebook)

        notebook.add(abonents, text="Абоненты")
        notebook.add(tariffs, text="Тарифы и Услуги")
        notebook.add(reports, text="Отчеты")
        notebook.add(settings, text="Настройки")

        # Наполнение вкладок
        self._setup_abonents_tab(abonents)
        self._setup_tariffs_tab(tariffs)
        self._setup_reports_tab(reports)
        self._setup_settings_tab(settings)

        self.date_todey = date.today()

        if 1 < self.date_todey.day < 10:
            self._accrual_of_amounts()

    def _setup_abonents_tab(self, frame):
        """Упрощенная вкладка 'Абоненты' с универсальным поиском и сортировкой"""

        search_frame = ttk.Frame(frame)
        search_frame.pack(fill="x", padx=10, pady=10)

        ttk.Label(search_frame, text="Поиск (Л/С, ФИО или Адрес):").pack(side="left", padx=5)

        self.search_entry = ttk.Entry(search_frame)
        self.search_entry.pack(side="left", padx=5, fill="x", expand=True)

        self.search_entry.bind("<Return>", lambda e: self._search_clients())

        ttk.Button(search_frame, text="Найти", command=self._search_clients).pack(side="left", padx=5)
        ttk.Button(search_frame, text="Сброс", command=self._load_clients).pack(side="left", padx=5)

        tree_container = ttk.Frame(frame)
        tree_container.pack(fill="both", expand=True, padx=10, pady=5)

        columns = ("account", "fio", "address", "tariff", "balance", "status")
        self.client_tree = ttk.Treeview(tree_container, columns=columns, show='headings')

        scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.client_tree.yview)
        self.client_tree.configure(yscrollcommand=scrollbar.set)

        self.client_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        headers = {
            "account": "Лицевой счет",
            "fio": "ФИО",
            "address": "Адрес",
            "tariff": "Тариф",
            "balance": "Баланс",
            "status": "Статус"
        }

        for col_id, col_text in headers.items():
            self.client_tree.heading(col_id, text=col_text,
                                     command=lambda c=col_id: self._sort_column(c, False))

            width = 150 if col_id in ["fio", "address"] else 100
            self.client_tree.column(col_id, width=width, anchor="center" if width == 100 else "w")

        self.group_operations_frame = ttk.LabelFrame(frame, text="Управление")
        self.group_operations_frame.pack(fill="x", padx=10, pady=10)

        btns_subframe = ttk.Frame(self.group_operations_frame)
        btns_subframe.pack(fill="x", padx=5, pady=5)

        ttk.Button(btns_subframe, text="Внести оплату", command=self._add_payment).pack(side="left", padx=5)


        ttk.Separator(btns_subframe, orient="vertical").pack(side="left", fill="y", padx=10)

        ttk.Button(btns_subframe, text="Добавить", command=self._add_client).pack(side="left", padx=5)
        ttk.Button(btns_subframe, text="Изменить", command=self._open_edit_window).pack(side="left", padx=5)
        ttk.Button(btns_subframe, text="Удалить", command=self._delete_client).pack(side="left", padx=5)

        self.client_tree.bind("<Double-Button-1>", self._open_edit_window)

        self._load_clients()

    def _sort_column(self, col, reverse):
        """Сортировка содержимого Treeview"""
        data = [(self.client_tree.set(child, col), child) for child in self.client_tree.get_children('')]

        if col in ["balance", "account"]:
            data.sort(key=lambda x: float(x[0].replace(' ', '') or 0), reverse=reverse)
        else:
            data.sort(key=lambda x: x[0].lower(), reverse=reverse)

        for index, (val, child) in enumerate(data):
            self.client_tree.move(child, '', index)

        self.client_tree.heading(col, command=lambda: self._sort_column(col, not reverse))

    def _setup_tariffs_tab(self, frame):
        """Создает таблицы 'Тарифы' и 'Услуги'"""

        # --- СЕКЦИЯ 1: ТАРИФЫ ---
        tariffs_label_frame = ttk.LabelFrame(frame, text="Список основных тарифов", padding=10)
        tariffs_label_frame.pack(fill="both", expand=True, padx=10, pady=5)

        tree_container_t = ttk.Frame(tariffs_label_frame)
        tree_container_t.pack(fill="both", expand=True)

        self.tariffs_tree = ttk.Treeview(
            tree_container_t,
            columns=("name", "price", "status"),
            show='headings',
            height=6  # Ограничиваем начальную высоту
        )

        scrollbar_t = ttk.Scrollbar(tree_container_t, orient="vertical", command=self.tariffs_tree.yview)
        self.tariffs_tree.configure(yscrollcommand=scrollbar_t.set)
        self.tariffs_tree.pack(side="left", fill="both", expand=True)
        scrollbar_t.pack(side="right", fill="y")

        # Настройка колонок тарифов
        t_cols = {"name": ("Наименование тарифа", 250), "price": ("Цена/мес", 100), "status": ("Статус", 100)}
        for cid, (txt, width) in t_cols.items():
            self.tariffs_tree.heading(cid, text=txt,
                                      command=lambda c=cid: self._sort_column(self.tariffs_tree, c, False))
            self.tariffs_tree.column(cid, anchor="center", width=width)

        # Кнопки тарифов
        t_btns = ttk.Frame(tariffs_label_frame)
        t_btns.pack(fill="x", pady=5)
        ttk.Button(t_btns, text="Добавить тариф", command=self._add_tariff).pack(side="left", padx=5)
        ttk.Button(t_btns, text="Удалить тариф", command=self._delete_tariff).pack(side="left", padx=5)

        # --- СЕКЦИЯ 2: УСЛУГИ ---
        services_label_frame = ttk.LabelFrame(frame, text="Услуги", padding=10)
        services_label_frame.pack(fill="both", expand=True, padx=10, pady=5)

        tree_container_s = ttk.Frame(services_label_frame)
        tree_container_s.pack(fill="both", expand=True)

        self.services_tree = ttk.Treeview(
            tree_container_s,
            columns=("name", "cost"),
            show='headings',
            height=6
        )

        scrollbar_s = ttk.Scrollbar(tree_container_s, orient="vertical", command=self.services_tree.yview)
        self.services_tree.configure(yscrollcommand=scrollbar_s.set)
        self.services_tree.pack(side="left", fill="both", expand=True)
        scrollbar_s.pack(side="right", fill="y")

        # Настройка колонок услуг
        s_cols = {"name": ("Наименование услуги", 250), "cost": ("Стоимость", 100)}
        for cid, (txt, width) in s_cols.items():
            self.services_tree.heading(cid, text=txt,
                                       command=lambda c=cid: self._sort_column(self.services_tree, c, False))
            self.services_tree.column(cid, anchor="center", width=width)

        # Кнопки услуг
        s_btns = ttk.Frame(services_label_frame)
        s_btns.pack(fill="x", pady=5)
        ttk.Button(s_btns, text="Добавить услугу", command=self._add_service).pack(side="left", padx=5)
        ttk.Button(s_btns, text="Удалить услугу", command=self._delete_service).pack(side="left", padx=5)

        # Загрузка данных
        self._load_tariffs()
        self._load_services()

    def _setup_reports_tab(self, frame):
        """Создает элементы для вкладки 'Отчеты'"""
        frame.columnconfigure(0, weight=1)
        current_row = 0

        abonents_report_frame = ttk.LabelFrame(frame, text="Отчеты по абонентам")
        abonents_report_frame.grid(row=current_row, column=0, sticky='we', padx=5, pady=10)
        abonents_report_frame.columnconfigure(0, weight=1)
        abonents_report_frame.rowconfigure(0, weight=1)
        current_row += 1

        buttons_frame_abonents = ttk.Frame(abonents_report_frame)
        buttons_frame_abonents.grid(row=current_row, column=0, sticky='ew',
                                    pady=15)
        ttk.Button(buttons_frame_abonents, text="Список должников", command=self._get_debtors_clients).pack(side="left",
                                                                                                            padx=5)
        ttk.Button(buttons_frame_abonents, text="Список абонентов по домам").pack(side="left", padx=5)
        ttk.Separator(
            buttons_frame_abonents,
            orient="vertical"
        ).pack(side="left", fill="y", padx=10, pady=5)

        self.reports_analysis_box = ttk.Combobox(buttons_frame_abonents, state="readonly",
                                                 values=["Количество подключений", "Количество отключений",
                                                         "Количество приостановленных"],
                                                 width=25)
        self.reports_analysis_box.pack(side="left", padx=5, pady=5)
        self.reports_analysis_box.current(0)

        ttk.Button(buttons_frame_abonents, text="Месячный отчет", command=self._get_result_report_analysis).pack(
            side="left", padx=5)
        ttk.Label(buttons_frame_abonents, text="Отчет: ").pack(side="left", padx=5, pady=5)
        self.result_report_analysis_lebel = (ttk.Label(buttons_frame_abonents))
        self.result_report_analysis_lebel.pack(side="left", padx=5, pady=5)
        current_row += 1

        analytical_reports_frame = ttk.LabelFrame(frame, text="Аналитические отчеты")
        analytical_reports_frame.grid(row=current_row, column=0, sticky='ew', padx=5, pady=10)
        analytical_reports_frame.columnconfigure(0, weight=1)
        current_row += 1

        buttons_frame_analytical_reports = ttk.Frame(analytical_reports_frame)
        buttons_frame_analytical_reports.grid(row=0, column=0, sticky='ew', pady=10)

        ttk.Label(buttons_frame_analytical_reports, text="Доходы за период").pack(side="left", padx=5)

        # Дата начала
        ttk.Label(buttons_frame_analytical_reports, text="с").pack(side="left", padx=5)
        self.start_date = DateEntry(buttons_frame_analytical_reports, width=12, background='darkblue',
                                    foreground='white', borderwidth=2,
                                    date_pattern='dd.mm.yyyy')
        self.start_date.pack(side="left", padx=5)

        # Дата конца
        ttk.Label(buttons_frame_analytical_reports, text="по").pack(side="left", padx=5)
        self.end_date = DateEntry(buttons_frame_analytical_reports, width=12, background='darkblue',
                                  foreground='white', borderwidth=2,
                                  date_pattern='dd.mm.yyyy')
        self.end_date.pack(side="left", padx=5)

        # Привязываем событие изменения даты начала для проверки
        self.start_date.bind("<<DateEntrySelected>>", self._validate_dates)
        self.end_date.bind("<<DateEntrySelected>>", self._validate_dates)

        ttk.Button(
            buttons_frame_analytical_reports,
            text="Сформировать",
            command=self._get_reports_income_for_period
        ).pack(side="left", padx=5)

        downloading_reports_frame = ttk.LabelFrame(frame, text="Работа с банком")
        downloading_reports_frame.grid(row=current_row, column=0, sticky='ew', padx=5, pady=10)
        downloading_reports_frame.columnconfigure(0, weight=1)
        downloading_reports_frame.rowconfigure(0, weight=1)
        current_row += 1

        buttons_frame_downloading_reports = ttk.Frame(downloading_reports_frame)
        buttons_frame_downloading_reports.grid(row=current_row, column=0, sticky='ew', padx=5, pady=10)
        ttk.Button(buttons_frame_downloading_reports, text="Выгрузить реестр для банка",
                   command=self._get_report_for_bank).pack(side="left", padx=5)
        ttk.Button(buttons_frame_downloading_reports, text="Загрузить реестр из банка",
                   command=self._set_report_for_bank).pack(side="left", padx=5)
        current_row += 1

    def _validate_dates(self, event=None):
        """Проверка, чтобы дата начала была не позже даты конца"""
        start = self.start_date.get_date()
        end = self.end_date.get_date()

        if start > end:
            self.end_date.set_date(start)

    def _get_reports_income_for_period(self):
        """Формирование дохода за период."""
        start_date = self.start_date.get_date()
        end_date = self.end_date.get_date()
        actual_end = datetime.combine(end_date, time.max)

        window_report = WindowReport(self,
                                     f"Список платежей за период с {start_date.strftime("%d.%m.%Y")} по {end_date.strftime("%d.%m.%Y")}",
                                     1, start_date, actual_end)

    def _search_clients(self):
        """Выполняет поиск клиентов (синхронная версия)."""
        val = self.search_entry.get().strip()
        if not val:
            return self._load_clients()
        clients = None
        for db in get_db():
            clients = search_clients(db, val)

        self._display_clients(clients)
        return None

    def _load_clients(self):
        """Загружает и отображает список всех клиентов (или сброс поиска)."""
        # 1. Очистка Treeview
        for item in self.client_tree.get_children():
            self.client_tree.delete(item)

        clients = None

        # 2. Получение данных
        for db in get_db():
            clients = get_clients(db)
            break

        # 3. Отображение
        self._display_clients(clients)

    def _display_clients(self, clients):
        """Отображает список объектов клиентов в Treeview."""
        # Очистка Treeview
        for item in self.client_tree.get_children():
            self.client_tree.delete(item)

        for client in clients:
            # Преобразуем объект SQLAlchemy в удобный кортеж
            is_active_status = "Да" if client.is_active == 1 else "Нет"

            self.client_tree.insert("", "end", values=(
                client.personal_account,
                client.full_name,
                client.address,
                client.tariff,
                f"{client.balance:.2f}",  # Форматируем баланс
                client.status.value,
            ))

    def _display_tariffs(self, tariffs):
        """Отображает список объектов тарифов в Treeview."""
        # Очистка Treeview
        for item in self.tariffs_tree.get_children():
            self.tariffs_tree.delete(item)

        for tariff in tariffs:
            # Преобразуем объект SQLAlchemy в удобный кортеж
            is_active_status = "Да" if tariff.is_active == 1 else "Нет"

            self.tariffs_tree.insert("", "end", values=(
                tariff.name,
                tariff.monthly_price,
                is_active_status
            ))

    def _display_services(self, services):
        """Отображает список объектов тарифов в Treeview."""
        # Очистка Treeview
        for item in self.services_tree.get_children():
            self.services_tree.delete(item)

        for service in services:
            self.services_tree.insert("", "end", values=(
                service.service_name,
                service.service_price,
            ))

    def _setup_tariff_tab(self, frame):

        # Виджет Treeview для отображения данных
        self.tariffs_tree = ttk.Treeview(
            frame,
            columns=("Название", "Стоимость"),
            show='headings'  # Скрываем первый столбец с индексами
        )
        self.tariffs_tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Настройка заголовков столбцов
        for col in self.tariffs_tree['columns']:
            self.tariffs_tree.heading(col, text=col)
            self.tariffs_tree.column(col, anchor="w", width=80)

        # Загрузка данных при старте
        self._load_tariffs()

    def _load_tariffs(self):
        """Загружает и отображает список всех Тарифов."""
        # 1. Очистка Treeview
        for item in self.tariffs_tree.get_children():
            self.tariffs_tree.delete(item)
        tariffs = None
        # 2. Получение данных
        for db in get_db():
            tariffs = get_tariffs(db)
            break

        # 3. Отображение
        self._display_tariffs(tariffs)

    def _load_services(self):
        """Загружает и отображает список всех Услуг."""
        # 1. Очистка Treeview
        for item in self.services_tree.get_children():
            self.services_tree.delete(item)
        service = None
        # 2. Получение данных
        for db in get_db():
            service = get_services(db)
            break

        # 3. Отображение
        self._display_services(service)

    def _delete_client(self):
        """Обрабатывает нажатие кнопки 'Удалить клиента'."""
        select_client = self.client_tree.item(self.client_tree.focus()).get('values')
        if not select_client:
            messagebox.showerror(
                "Внимание!",
                "Необходимо выбрать абонента!"
            )
        else:
            try:
                for db in get_db():
                    client = get_client_by_pa(db, select_client[0])
                    delete_yes = messagebox.askyesno(
                        "Внимание!",
                        f"Вы действительно хотите удалить абонента {client.full_name}, ЛС: {client.personal_account}?"
                    )
                    if delete_yes:
                        delete_client(db, int(client.id))
                        messagebox.showinfo(
                            "Успех",
                            f"Клиент {client.full_name} (ID: {client.id}) успешно удален!"
                        )
                    break
            except Exception as e:
                messagebox.showerror("Ошибка удаления", f"Не удалось удалить клиента:\n{e}")

            self._search_clients()

    def _edit_client(self):
        """Обрабатывает нажатие кнопки 'Редактировать клиента'."""
        client_id = None
        select_client = self.client_tree.item(self.client_tree.focus()).get('values')
        if not select_client:
            messagebox.showerror(
                "Внимание!",
                "Необходимо выбрать абонента!"
            )
        else:
            client_personal_account = select_client[0]

            try:
                if client_personal_account:
                    for db in get_db():
                        client = get_client_by_pa(db, client_personal_account)
                        current_client = ClientBase(
                            personal_account=int(client.personal_account),
                            full_name=str(client.full_name),
                            address=str(client.address),
                            phone_number=str(client.phone_number),
                            tariff=str(client.tariff),
                            balance=float(client.balance),
                        )
                        new_window_edit_client = WindowAddClient(self)
                        new_window_edit_client.set_data_client(current_client)
                        break


            except Exception as e:
                print(e)

    def _delete_tariff(self):
        """Обрабатывает нажатие кнопки 'Удалить Тариф'."""
        select_tariff = self.tariffs_tree.item(self.tariffs_tree.focus()).get('values')
        if not select_tariff:
            messagebox.showerror(
                "Внимание!",
                "Необходимо выбрать тариф!"
            )
        else:
            try:
                for db in get_db():
                    tariff = get_tariff_by_name(db, select_tariff[0])
                    delete_tariff(db, int(tariff.id))
                    messagebox.showinfo(
                        "Успех",
                        f"Тариф {tariff.name} успешно удален!"
                    )
                    break
            except Exception as e:
                messagebox.showerror("Ошибка удаления", f"Не удалось удалить тариф:\n{e}")

            self._load_tariffs()

    def _delete_service(self):
        """Обрабатывает нажатие кнопки 'Удалить Услугу'."""
        select_service = self.services_tree.item(self.services_tree.focus()).get('values')
        if not select_service:
            messagebox.showerror(
                "Внимание!",
                "Необходимо выбрать Услугу!"
            )
        else:
            try:
                for db in get_db():
                    service = get_service_by_name(db, select_service[0])
                    delete_service(db, int(service.id))
                    messagebox.showinfo(
                        "Успех",
                        f"Услуга {service.service_name} успешно удалена!"
                    )
                    break
            except Exception as e:
                messagebox.showerror("Ошибка удаления", f"Не удалось удалить Услугу:\n{e}")

            self._load_services()

    def _add_client(self):
        """Создание нового окна для добавления клиента."""

        add_window = WindowAddClient(self)

    def _add_payment(self):
        """Создание окна для внесения оплаты."""
        # TODO
        client_id = None
        select_client = self.client_tree.item(self.client_tree.focus()).get('values')
        if not select_client:
            messagebox.showerror(
                "Внимание!",
                "Необходимо выбрать абонента!"
            )
        else:
            client_personal_account = select_client[0]
            try:
                if client_personal_account:
                    for db in get_db():
                        client = get_client_by_pa(db, client_personal_account)
                        current_client = ClientForPayments(
                            personal_account=int(client.personal_account),
                            full_name=str(client.full_name),
                            address=str(client.address),
                            phone_number=str(client.phone_number),
                            tariff=str(client.tariff),
                            balance=float(client.balance),
                            is_active=bool(client.is_active),
                        )
                        new_window_edit_client = WindowAddPayment(self)
                        new_window_edit_client.set_data_client(current_client)
                        break


            except Exception as e:
                messagebox.showerror(
                    title="Ошибка!",
                    message=f"Произошла ошибка \nПодробнее: {e}"
                )

    def _add_tariff(self):
        """Создание нового окна для добавления тарифа."""
        add_window_tariff = WindowAddTariff(self)

    def _add_service(self):
        """Создание нового окна для добавления Услуги."""
        add_window_service = WindowAddService(self)

    def _set_client_satus(self):
        """Обрабатывает нажатие кнопки 'Приостановить'."""
        status = self.status_box.get()
        select_client = self.client_tree.item(self.client_tree.focus()).get('values')
        if not select_client:
            messagebox.showerror(
                "Внимание!",
                "Необходимо выбрать абонента!"
            )
        else:
            try:
                for db in get_db():
                    client = get_client_by_pa(db, select_client[0])
                    result = messagebox.askyesno(
                        "Подтверждение действия",
                        f"Вы уверены, что хотите изменить статус Абонента - {client.full_name} c '{client.status.value}' на '{status}'?"
                    )
                    if result:
                        set_client_status(db, int(client.id), status)
                        if status == StatusClientEnum.PAUSE or status == StatusClientEnum.DISCONNECTING:
                            set_client_activity(db, int(client.id), False)
                        elif status == StatusClientEnum.CONNECTING:
                            set_client_activity(db, int(client.id), True)
                    break
            except Exception as e:
                messagebox.showerror("Ошибка операции!", f"Не удалось изменить статус абонента! \nПодробности:\n{e}")

            self._search_clients()

    def _open_edit_window(self, event=None):
        """"""
        if event:
            item_id = self.client_tree.identify_row(event.y)
        else:
            selection = self.client_tree.selection()
            item_id = selection[0] if selection else None

        if not item_id:
            return

        item_data = self.client_tree.item(item_id)
        values = item_data['values']

        current_client = None
        try:
            for db in get_db():
                client = get_client_by_pa(db, values[0])
                current_client = ClientCard(
                    personal_account=client.personal_account,
                    full_name=client.full_name,
                    address=client.address,
                    phone_number=client.phone_number,
                    tariff=client.tariff,
                    balance=float(client.balance),
                    client_id=client.id,
                    is_active=client.is_active,
                    connection_date=client.connection_date,
                    passport=client.passport,
                    status=client.status,
                    status_date=client.status_date,
                )
                break
        except Exception as e:
            messagebox.showerror(
                "Ошибка!",
                f"Произошла ошибка!\nПодробнее:\n{e}"
            )
        new_window = WindowEditAndViewClient(self)
        new_window.set_data_client(current_client)
        # self._load_clients()

    def _accrual_of_amounts(self):
        """Метод начисления ежемесячной оплаты Абонентам."""
        db = next(get_db())
        try:
            clients = get_clients(db)
            today = self.date_todey

            for client in clients:
                # 1. Проверяем, было ли уже начисление (чтобы не начислить дважды за месяц)
                last_accrual = get_last_accrual_by_client(db, client.id)

                if last_accrual is None:
                    # Начисление оплаты клиенту, если клиент имеет статус Подключен
                    if client.status == StatusClientEnum.CONNECTING:
                        if client.status_date.month != today.month:
                            conn_date = client.connection_date

                            # Если подключение было в ПРОШЛОМ месяце или раньше
                            if conn_date.month != today.month or conn_date.year != today.year:
                                if apply_monthly_charge(db, client.id):
                                    client.accrual_date = today
                                    tariff = get_tariff_by_name(db, client.tariff)
                                    create_accrual_monthly(db, client, tariff, today)

                            # Если подключение в ЭТОМ месяце (пропорциональное начисление)
                            else:
                                # Узнаем сколько всего дней в месяце (для 2026 года)
                                _, days_in_month = calendar.monthrange(conn_date.year, conn_date.month)

                                # Считаем количество дней пользования: (Всего - ДеньПодключения + 1)
                                # Если подключился 11-го, то (31 - 11 + 1) = 21 день владения
                                actual_days = days_in_month - conn_date.day + 1

                                if apply_daily_charge(db, client.id, actual_days):
                                    client.accrual_date = today
                                    create_accrual_daily(db, client.id, actual_days, today)
                        elif client.status_date.month == today.month:
                            status_date = client.status_date
                            _, days_in_month = calendar.monthrange(status_date.year, status_date.month)

                            actual_days = days_in_month - conn_date.day + 1

                            if apply_daily_charge(db, client.id, actual_days):
                                client.accrual_date = today
                                create_accrual_daily(db, client.id, actual_days, today)

                    # Начисление оплаты клиенту, если клиент в текущем месяце был приостановлен
                    elif client.status == StatusClientEnum.PAUSE:
                        status_date = client.status_date
                        if status_date.month == today.month and status_date.year == today.year:
                            actual_days = status_date.day - 1

                            if apply_daily_charge(db, client.id, actual_days):
                                client.accrual_date = today
                                create_accrual_daily(db, client.id, actual_days, today)



            db.commit()  # Фиксируем все начисления одной транзакцией

        except Exception as e:
            db.rollback()
            messagebox.showerror("Ошибка!", f"Ошибка начисления оплаты!\n{e}")
        finally:
            db.close()

    def _get_debtors_clients(self):
        """Создание нового окна для отчета."""
        window_report = WindowReport(self, "Список должников", 0)

    def _get_result_report_analysis(self):
        """Метод получения месячного отчета по движению абонентов за месяц
        Вариант отчета: 'Количество подключений', 'Количество отключений', 'Количество приостановленных'
        """
        current_month = date.today().month
        type_report = self.reports_analysis_box.get()

        clients = None
        for db in get_db():
            clients = get_clients(db)
            break

        count_result = 0

        if type_report == "Количество подключений":
            for client in clients:
                if client.connection_date.month == current_month:
                    count_result += 1
            self.result_report_analysis_lebel.config(text=count_result)
        elif type_report == "Количество отключений":
            for client in clients:
                if client.status_date and client.status_date.month == current_month and client.status == StatusClientEnum.DISCONNECTING:
                    count_result += 1
            self.result_report_analysis_lebel.config(text=count_result)
        elif type_report == "Количество приостановленных":
            for client in clients:
                if client.status_date and client.status_date.month == current_month and client.status == StatusClientEnum.PAUSE:
                    count_result += 1
            self.result_report_analysis_lebel.config(text=count_result)

    def _get_report_for_bank(self):
        """
        Метод создает в папке 'reports' текстовый файл с именем '1117005066_40702810728180100104_дата_формирования.txt'
        Файл содержит данные о внесении оплаты в формате - 'номер_ЛС;Фамилия ИО;;ТВ;;сумма_платежа'. Каждый клиент на новой строчке
        :return:
        """
        clients = None
        current_month = date.today().month
        dir_path = Path("reports")
        file_path = dir_path / f"1117005066_40702810728180100104_{date.today()}.txt"
        dir_path.mkdir(parents=True, exist_ok=True)

        for db in get_db():
            clients = get_clients(db)
            break

        if clients:
            with file_path.open(mode="w", encoding="utf-8") as file:
                for client in clients:
                    personal_account = client.personal_account
                    client_full_name = client.full_name.split()
                    full_name = None
                    if len(client_full_name) == 3:
                        full_name = f"{client_full_name[0].capitalize()} {client_full_name[1][0].upper()}{client_full_name[2][0].upper()}"
                    else:
                        full_name = client_full_name[0].capitalize()
                    amount = self._get_last_payment_client(client.id, current_month)
                    record = f"{personal_account};{full_name};;ТВ;;{amount:.2f}\n"
                    if amount != 0:
                        file.write(record)
                    else:
                        continue

    def _set_report_for_bank(self):
        """Метод для загрузки данных с банка"""
        pass

    def _get_last_payment_client(self, client_id: int, current_month: int) -> float:
        result = 0
        for db in get_db():
            last_payment_client = get_last_payment_by_client(db, client_id)
            if last_payment_client and last_payment_client.payment_date.month == current_month:
                result = result + last_payment_client.amount
        return result

    def _setup_settings_tab(self, frame):
        """Создает элементы для вкладки 'Настройки'"""
        frame.columnconfigure(0, weight=1)
        current_row = 0

        db_frame = ttk.LabelFrame(frame, text="Действия с базой")
        db_frame.grid(row=current_row, column=0, sticky='we', padx=5, pady=10)
        db_frame.columnconfigure(0, weight=1)
        db_frame.rowconfigure(0, weight=1)
        current_row += 1

        buttons_frame = ttk.Frame(db_frame)
        buttons_frame.grid(row=current_row, column=0, sticky='ew', pady=15)
        ttk.Button(buttons_frame, text="Загрузить базу", command=self._select_file).pack(side="left", padx=5)
        ttk.Button(buttons_frame, text="Выгрузить базу", command=self._save_db_to_file).pack(side="left", padx=5)

        abonents_frame = ttk.LabelFrame(frame, text="Действия с абонентами")
        abonents_frame.grid(row=current_row, column=0, sticky='we', padx=5, pady=10)
        abonents_frame.columnconfigure(0, weight=1)
        abonents_frame.rowconfigure(0, weight=1)
        current_row += 1

        buttons_abonents_frame = ttk.Frame(abonents_frame)
        buttons_abonents_frame.grid(row=current_row, column=0, sticky='ew', pady=15)
        ttk.Button(buttons_abonents_frame, text="Выполнить ручное начисление", command=self._accrual_of_amounts).pack(
            side="left", padx=5)

    def _select_file(self):
        """
        Метод для загрузки файлов в программу.
        Структура файла: ЛС;ФИО;Адрес;Телефон;Тариф;Дата подключения;Баланс
        :return: # TODO Необходимо разделить логику загрузки данных в программу в отдельный метод.
        """
        ask_result = messagebox.askyesno(
            title="Подтверждение",
            message="Вы действительно хотите удалить существующую базу клиентов и загрузить новую?"
        )
        if ask_result:

            clients = []
            filename = filedialog.askopenfile()
            if filename:
                with filename:
                    list_lines_of_file = filename.readlines()
                    for line in list_lines_of_file:
                        clean_line = line.strip()
                        if not clean_line: continue
                        data_line = clean_line.split(";")
                        try:
                            client_data = ClientCreate(
                                personal_account=int(data_line[0]),
                                full_name=data_line[1],
                                address=data_line[2],
                                phone_number=data_line[3],
                                tariff=data_line[4],
                                connection_date=date.strptime(data_line[5], "%Y-%m-%d"),
                                balance=float(data_line[6].replace(',', '.')),
                            )
                            clients.append(client_data)
                        except (ValueError, IndexError) as e:
                            messagebox.showerror(
                                title="Ошибка!",
                                message=f"Ошибка в строке: {line}. Ошибка: {e}"
                            )
                            continue
                    db = next(get_db())
                    try:
                        clear_db_clients(db)
                        bulk_create_clients(db, clients)
                    finally:
                        db.close()
        messagebox.showinfo(
            title="Успешно!",
            message=f"База успешна обновлена, загружено {len(clients)} абонентов"
        )

    def _save_db_to_file(self):
        """Метод выгрузки данных абонентов в формате 'csv'.
        Файл 'data_clients_дата_выгрузки.csv' сохраняется в папке 'out', если папка отсутствует, метод ее создает.
        Структура содержимого файла: ЛС;ФИО;Адрес;Телефон;Тариф;Дата подключения;Баланс;Статус
        :return:
        """

        clients = None
        dir_path = Path("out")
        file_path = dir_path / f"data_clients_{date.today()}.csv"
        dir_path.mkdir(parents=True, exist_ok=True)

        for db in get_db():
            clients = get_clients(db)
            break

        if clients:
            with file_path.open(mode="w", encoding="utf-8") as file:
                for client in clients:
                    record = f"{client.personal_account};{client.full_name};{client.address};{client.phone_number};{client.tariff};{client.connection_date.strftime("%Y-%m-%d")};{client.balance:.2f};{client.status.value}\n"
                    if record:
                        file.write(record)
                    else:
                        continue
            messagebox.showinfo(
                title="Успешно!",
                message=f"Данные абонентов выгружены в папку '{dir_path}'!"
            )


class WindowAddClient(tkinter.Toplevel):
    """Класс для вызова окна добавления клиента."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Добавить клиента')
        self.geometry('400x300')
        self.resizable(False, False)

        # Заголовки и поля ввода
        ttk.Label(self, text="Лицевой счет:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.personal_account_entry = ttk.Entry(self, width=40)
        self.personal_account_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(self, text="ФИО:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.full_name_entry = ttk.Entry(self, width=40)
        self.full_name_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(self, text="Адрес:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.address_entry = ttk.Entry(self, width=40)
        self.address_entry.grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(self, text="Телефон:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.phone_entry = ttk.Entry(self, width=40)
        self.phone_entry.grid(row=3, column=1, padx=5, pady=5)

        ttk.Label(self, text="Тариф:").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.tariff_entry = ttk.Combobox(self, state="readonly", width=37)
        self.tariff_entry["values"] = (self._get_tariffs() if len(self._get_tariffs()) > 0 else ["Нет тарифов"])
        self.tariff_entry.current(0)
        self.tariff_entry.grid(row=4, column=1, padx=5, pady=5)

        ttk.Label(self, text="Баланс:").grid(row=5, column=0, padx=5, pady=5, sticky="w")
        self.balance_entry = ttk.Entry(self, width=40)
        self.balance_entry.grid(row=5, column=1, padx=5, pady=5)

        ttk.Label(self, text="Дата подключения:").grid(row=6, column=0, padx=5, pady=5, sticky="w")
        self.cal_entry = DateEntry(self, width=37, date_pattern="dd.mm.yyyy")
        self.cal_entry.grid(row=6, column=1, padx=5, pady=5)

        # Кнопка добавления
        ttk.Button(self, text="Отправить", command=self._send_data_client).grid(
            row=7, column=0, columnspan=1, pady=10
        )
        ttk.Button(self, text="Очистить поля", command=self._clear_add_fields).grid(
            row=7, column=1, columnspan=1, pady=10
        )

        self.transient(parent)
        self.grab_set()

    def _add_client(self, data: ClientCreate):
        """Обрабатывает нажатие кнопки "Добавить Клиента"."""
        try:

            # 2. Вызов синхронной CRUD-функции
            for db in get_db():
                new_client = create_client(db, data)
                if new_client:
                    messagebox.showinfo(
                        "Успех",
                        f"Клиент {new_client.full_name} (ID: {new_client.id}) успешно добавлен!"
                    )
                break

            self.destroy()

        except Exception as e:
            messagebox.showerror("Ошибка добавления", f"Не удалось добавить клиента:\n{e}")

    def _update_client(self, client_id: int, client: ClientUpdate):
        """Редактирование Клиента в дополнительном окне.

        :param client_id: ID клиента в базе данных.
        :param client: Данные выбранного Клиента в главном окне.
        """
        try:

            # 2. Вызов синхронной CRUD-функции
            for db in get_db():
                current_client = update_client(db, client_id, client)
                if current_client:
                    messagebox.showinfo(
                        "Успех",
                        f"Клиент {current_client.full_name} (ID: {current_client.id}) успешно изменен!"
                    )
                break

            self.destroy()

        except Exception as e:
            messagebox.showerror("Ошибка изменения", f"Не удалось изменить клиента:\n{e}")

    def _clear_add_fields(self):
        """Очищает поля ввода после добавления."""

        self.personal_account_entry.delete(0, END)
        self.full_name_entry.delete(0, END)
        self.address_entry.delete(0, END)
        self.phone_entry.delete(0, END)
        self.tariff_entry.delete(0, END)
        self.balance_entry.delete(0, END)
        self.balance_entry.delete(0, END)

    def set_data_client(self, client: ClientBase):
        """Заполняем данные клиента с базы для редактирования"""
        self.personal_account_entry.insert(0, client.personal_account)
        self.full_name_entry.insert(0, client.full_name)
        self.address_entry.insert(0, client.address)
        self.phone_entry.insert(0, client.phone_number)
        self.balance_entry.insert(0, client.balance)

    def _send_data_client(self):
        """Общий метод для реализации добавления клиента в базу
         или обновления клиента в базе."""
        # 1. Сбор данных
        window_data = {
            "personal_account": self.personal_account_entry.get(),
            "full_name": self.full_name_entry.get(),
            "address": self.address_entry.get(),
            "phone_number": self.phone_entry.get(),
            "tariff": self.tariff_entry.get(),
            "balance": self.balance_entry.get(),
            "connection_date": self.cal_entry.get_date(),
        }

        if window_data.values():
            try:
                data = ClientCreate(**window_data)
                for db in get_db():
                    # Проверка клиента в базе
                    client = get_client_by_pa(db, int(data.personal_account))
                    if not client:
                        self._add_client(data)
                    else:
                        current_client = ClientUpdate(
                            full_name=self.full_name_entry.get(),
                            address=self.address_entry.get(),
                            phone_number=self.phone_entry.get(),
                            tariff=self.tariff_entry.get(),
                        )
                        self._update_client(int(client.id), current_client)

            except (Exception, ValidationError) as e:
                messagebox.showerror(
                    "Ошибка",
                    f"Возникла ошибка!\nПодробности: {e}"
                )
        else:
            messagebox.showerror(
                "Ошибка!",
                "Заполните все поля!"
            )

    def _get_tariffs(self):
        """Получает тарифы из базы и возвращает списком"""
        try:
            list_tariffs = []
            for db in get_db():
                tariffs = get_tariffs(db)
                if tariffs:

                    for tariff in tariffs:
                        list_tariffs.append(tariff.name)
                    return list_tariffs
                else:
                    messagebox.showerror(
                        "Ошибка!",
                        "Нет тарифов, сначала необходимо добавить тариф!"
                    )
                    self.focus()
                break
            return list_tariffs
        except Exception as e:

            self.destroy()


class WindowAddPayment(tkinter.Toplevel):
    """Класс для вызова окна внесения оплаты."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Внести оплату")
        self.geometry("330x220")
        self.resizable(False, False)

        # Создаем фрейм (рамку) для лучшего размещения элементов (Padding)
        main_frame = ttk.Frame(self, padding="20 20 20 20")
        main_frame.pack(fill='both', expand=True)

        # Использование сетки (Grid) для расположения элементов

        ttk.Label(main_frame, text="ЛC Клиента:").grid(column=0, row=0, sticky=tkinter.W, pady=5)
        self.personal_account = tkinter.StringVar()
        personal_account_label = ttk.Label(main_frame, textvariable=self.personal_account, width=30)
        personal_account_label.grid(column=1, row=0, padx=10, pady=5, sticky=tkinter.E)

        ttk.Label(main_frame, text="ФИО Клиента:").grid(column=0, row=1, sticky=tkinter.W, pady=5)
        self.full_name_text = tkinter.StringVar()
        full_name_label = ttk.Label(main_frame, textvariable=self.full_name_text, width=30)
        full_name_label.grid(column=1, row=1, padx=10, pady=5, sticky=tkinter.E)

        ttk.Label(main_frame, text="Баланс клиента:").grid(column=0, row=2, sticky=tkinter.W, pady=5)
        self.balance_text = tkinter.StringVar()
        balance_label = ttk.Label(main_frame, textvariable=self.balance_text, width=30)
        balance_label.grid(column=1, row=2, padx=10, pady=5, sticky=tkinter.E)

        ttk.Label(main_frame, text="Статус клиента:").grid(column=0, row=3, sticky=tkinter.W, pady=5)
        self.status_text = tkinter.StringVar()
        status_label = ttk.Label(main_frame, textvariable=self.status_text, width=30)
        status_label.grid(column=1, row=3, padx=10, pady=5, sticky=tkinter.E)

        ttk.Label(main_frame, text="Сумма (RUB):").grid(column=0, row=4, sticky=tkinter.W, pady=5)
        self.amount_entry = ttk.Entry(main_frame, width=30)
        # 'W' (west) означает выравнивание по левому краю
        self.amount_entry.grid(column=1, row=4, padx=10, pady=5, sticky=tkinter.E)
        self.amount_entry.insert(0, "600.00")  # Значение по умолчанию

        # --- 4. Кнопка "Оплатить" ---
        payment_button = ttk.Button(main_frame, text="Оплатить", command=self._process_payment)
        # Размещаем кнопку на всю ширину под полями
        payment_button.grid(column=0, row=5, columnspan=2, pady=20, sticky=tkinter.W + tkinter.E)

        self.transient(parent)
        self.grab_set()

    def _process_payment(self):
        """Обработка нажатия кнопки добавления платежа."""
        try:

            personal_account = int(self.personal_account.get())
            amount = float(self.amount_entry.get())
            if personal_account and amount > 0.0:
                for db in get_db():
                    current_client = get_client_by_pa(db, personal_account)

                    new_payment = PaymentCreate(
                        amount=amount,
                        client_id=int(current_client.id),
                    )
                    new_payment_db = create_payment(db, new_payment)

                    new_balance = amount + current_client.balance
                    client_for_update = ClientUpdate(
                        balance=new_balance,
                    )
                    current_client.balance += amount
                    update_client(db, int(current_client.id), client_for_update)
                    messagebox.showinfo(
                        "Успех!",
                        f"Внесена сумма: {amount} руб. \nдля Клиента: {current_client.full_name} \nЛицевой счёт: {current_client.personal_account}"
                    )

                    break
            self.destroy()
        except (Exception, ValidationError) as e:
            messagebox.showerror(
                "Ошибка!",
                f"Произошла ошибка, подробнее: \n{e}"
            )

    def set_data_client(self, current_client: ClientForPayments):
        """Заполняем данные клиента с базы"""
        self.personal_account.set(str(current_client.personal_account))
        self.full_name_text.set(current_client.full_name)
        self.balance_text.set(str(current_client.balance))
        self.status_text.set("Активный" if current_client.is_active else "Приостановлен")


class WindowAddTariff(tkinter.Toplevel):
    """Класс для вызова окна добавления тарифа (Версия 2026)."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Добавить тариф')
        self.geometry('400x200')
        self.resizable(False, False)

        # Используем фрейм для управления версткой
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill="both", expand=True)

        ttk.Label(main_frame, text="Наименование тарифа:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.name_tariff = ttk.Entry(main_frame, width=30)
        self.name_tariff.grid(row=0, column=1, padx=5, pady=5, sticky="we")

        ttk.Label(main_frame, text="Стоимость в месяц:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.tariff_price_entry = ttk.Entry(main_frame, width=30)
        self.tariff_price_entry.grid(row=1, column=1, padx=5, pady=5, sticky="we")

        # Кнопка добавления
        ttk.Button(main_frame, text="Отправить", command=self._add_tariff).grid(
            row=2, column=0, columnspan=2, pady=15
        )

        # Растягиваем вторую колонку, чтобы поля ввода занимали все место
        main_frame.columnconfigure(1, weight=1)

        self._center_to_parent(parent)

        self.transient(parent)
        self.grab_set()

    def _center_to_parent(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (self.winfo_width() // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

    def _add_tariff(self):
        """Обрабатывает нажатие кнопки "Добавить тариф"."""
        name = self.name_tariff.get().strip()
        price_str = self.tariff_price_entry.get().strip()

        # 1. Валидация ввода
        if not name or not price_str:
            messagebox.showwarning("Внимание", "Заполните все поля.")
            return

        try:
            # Преобразуем цену в число (float)
            price = float(price_str.replace(',', '.'))
        except ValueError:
            messagebox.showerror("Ошибка ввода", "Некорректный формат цены. Используйте цифры и точку.")
            return

        try:
            window_data = {
                "name": name,
                "monthly_price": price,
            }
            if window_data:

                # 2. Вызов синхронной CRUD-функции
                for db in get_db():
                    new_tariff = create_tariff(db, TariffCreate(**window_data))
                    messagebox.showinfo(
                        "Успех",
                        f"Тариф {new_tariff.name} успешно добавлен!"
                    )
                    break
            self.destroy()

        except Exception as e:
            messagebox.showerror("Ошибка добавления", f"Не удалось добавить тариф:\n{e}")


class WindowAddService(tkinter.Toplevel):
    """Класс для вызова окна добавления Услуги."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Добавить Услугу')
        self.geometry('400x200')
        self.resizable(False, False)

        # Используем фрейм для управления версткой
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill="both", expand=True)

        ttk.Label(main_frame, text="Наименование Услуги:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.name_service = ttk.Entry(main_frame, width=30)
        self.name_service.grid(row=0, column=1, padx=5, pady=5, sticky="we")

        ttk.Label(main_frame, text="Стоимость в месяц:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.service_price_entry = ttk.Entry(main_frame, width=30)
        self.service_price_entry.grid(row=1, column=1, padx=5, pady=5, sticky="we")

        # Кнопка добавления
        ttk.Button(main_frame, text="Отправить", command=self._add_service).grid(
            row=2, column=0, columnspan=2, pady=15
        )

        # Растягиваем вторую колонку, чтобы поля ввода занимали все место
        main_frame.columnconfigure(1, weight=1)

        self._center_to_parent(parent)

        self.transient(parent)
        self.grab_set()

    def _center_to_parent(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (self.winfo_width() // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

    def _add_service(self):
        """Обрабатывает нажатие кнопки "Добавить тариф"."""
        name = self.name_service.get().strip()
        price_str = self.service_price_entry.get().strip()

        # 1. Валидация ввода
        if not name or not price_str:
            messagebox.showwarning("Внимание", "Заполните все поля.")
            return

        try:
            # Преобразуем цену в число (float)
            price = float(price_str.replace(',', '.'))
        except ValueError:
            messagebox.showerror("Ошибка ввода", "Некорректный формат цены. Используйте цифры и точку.")
            return

        try:
            window_data = {
                "service_name": name,
                "service_price": price,
            }
            if window_data:

                # 2. Вызов синхронной CRUD-функции
                for db in get_db():
                    new_service = create_service(db, ServiceCreate(**window_data))
                    messagebox.showinfo(
                        "Успех",
                        f"Услуга {new_service.service_name} успешно добавлена!"
                    )
                    break
            self.destroy()

        except Exception as e:
            messagebox.showerror("Ошибка добавления", f"Не удалось добавить услугу:\n{e}")


class WindowEditAndViewClient(tkinter.Toplevel):
    """Класс для вызова окна Карточка абонента."""

    def __init__(self, parent):
        super().__init__(parent)

        self.title("Карточка абонента")
        self.geometry("650x750")
        self.resizable(False, False)

        main_frame = ttk.Frame(self, padding=20)
        main_frame.pack(fill='both', expand=True)

        main_frame.columnconfigure(1, weight=1)

        current_row = 0

        ttk.Label(main_frame, text="Лицевой счет:").grid(row=current_row, column=0, sticky='w', padx=5, pady=5)
        self.personal_account_entry = ttk.Entry(main_frame)
        self.personal_account_entry.grid(row=current_row, column=1, sticky='we', padx=5, pady=5)
        current_row += 1

        ttk.Label(main_frame, text="ФИО:").grid(row=current_row, column=0, sticky='w', padx=5, pady=5)
        self.full_name_entry = ttk.Entry(main_frame)
        self.full_name_entry.grid(row=current_row, column=1, sticky='we', padx=5, pady=5)
        current_row += 1

        ttk.Label(main_frame, text="Адрес:").grid(row=current_row, column=0, sticky='nw', padx=5, pady=5)
        self.text_address = tkinter.Entry(main_frame, width=40)
        self.text_address.grid(row=current_row, column=1, sticky='we', padx=5, pady=5)
        current_row += 1

        ttk.Label(main_frame, text="Телефон:").grid(row=current_row, column=0, sticky='nw', padx=5, pady=5)
        self.phone_entry = ttk.Entry(main_frame, width=40)
        self.phone_entry.grid(row=current_row, column=1, sticky='we', padx=5, pady=5)
        current_row += 1

        ttk.Label(main_frame, text="Тариф:").grid(row=current_row, column=0, sticky='w', padx=5, pady=5)

        tariff_balance_subframe = ttk.Frame(main_frame)
        tariff_balance_subframe.grid(row=current_row, column=1, sticky='we')

        self.tariff_entry = ttk.Combobox(tariff_balance_subframe, state="readonly", values=self._get_tariffs(),
                                         width=20)
        self.tariff_entry.pack(side='left', padx=5)

        ttk.Label(tariff_balance_subframe, text="Баланс абонента:").pack(side='left', padx=5)
        self.balance_entry = ttk.Entry(tariff_balance_subframe, width=10)
        self.balance_entry.pack(side='left', padx=5)

        current_row += 1

        ttk.Label(main_frame, text="Статус:").grid(row=current_row, column=0, sticky='w', padx=5, pady=5)

        status_set_subframe = ttk.Frame(main_frame)
        status_set_subframe.grid(row=current_row, column=1, sticky='we')

        self.status_list = [StatusClientEnum.CONNECTING.value, StatusClientEnum.PAUSE.value,
                            StatusClientEnum.DISCONNECTING.value]
        self.combo_status = ttk.Combobox(status_set_subframe, state="readonly",
                                         values=self.status_list)
        self.combo_status.pack(side='left', padx=5)

        ttk.Label(status_set_subframe, text="Дата изменения статуса:").pack(side='left', padx=5)

        self.status_date_entry = DateEntry(status_set_subframe, date_pattern="dd.mm.yyyy")
        self.status_date_entry.pack(side='left', padx=5)
        current_row += 1

        ttk.Label(main_frame, text="Дата подключения:").grid(row=current_row, column=0, sticky='w', padx=5, pady=5)
        self.connect_date_entry = DateEntry(main_frame, date_pattern="dd.mm.yyyy")
        self.connect_date_entry.grid(row=current_row, column=1, sticky='w', padx=5, pady=5)  # sticky='w'
        current_row += 1

        passport_frame = ttk.LabelFrame(main_frame, text="Паспортные данные")
        passport_frame.grid(row=current_row, column=0, columnspan=2, sticky='we', padx=5, pady=10)

        p_row = 0  # Локальный счетчик для фрейма
        ttk.Label(passport_frame, text="Серия и номер:").grid(row=p_row, column=0, sticky='w', padx=5, pady=5)
        self.passport_ser_num = ttk.Entry(passport_frame)
        self.passport_ser_num.grid(row=p_row, column=1, sticky='we', padx=5, pady=5)
        passport_frame.columnconfigure(1, weight=1)  # Растягиваем поле ввода
        p_row += 1

        ttk.Label(passport_frame, text="Дата выдачи:").grid(row=current_row, column=0, sticky='w', padx=5, pady=5)
        self.passport_data = ttk.Entry(passport_frame)
        self.passport_data.grid(row=current_row, column=1, sticky='we', padx=5, pady=5)
        current_row += 1

        ttk.Label(passport_frame, text="Кем выдан:").grid(row=current_row, column=0, sticky='w', padx=5, pady=5)
        self.passport_how = ttk.Entry(passport_frame)
        self.passport_how.grid(row=current_row, column=1, sticky='we', padx=5, pady=5)
        current_row += 1

        accruals_frame = ttk.LabelFrame(main_frame, text="Список начислений")
        accruals_frame.grid(row=current_row, column=0, columnspan=2, sticky='we', padx=5, pady=10)

        accruals_frame.columnconfigure(0, weight=1)
        accruals_frame.rowconfigure(0, weight=1)

        cols = ('date', 'amount', 'per_month')
        self.tree_accruals = ttk.Treeview(accruals_frame, columns=cols, show='headings', height=5)

        # Настраиваем заголовки
        self.tree_accruals.heading('date', text='Дата')
        self.tree_accruals.heading('amount', text='Сумма')
        self.tree_accruals.heading('per_month', text='За месяц')

        # Настраиваем ширину колонок
        self.tree_accruals.column('date', width=100, anchor='center')
        self.tree_accruals.column('amount', width=100, anchor='center')
        self.tree_accruals.column('per_month', width=150, anchor='center')

        scrollbar = ttk.Scrollbar(accruals_frame, orient="vertical", command=self.tree_accruals.yview)
        self.tree_accruals.configure(yscrollcommand=scrollbar.set)

        self.tree_accruals.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')

        current_row += 1

        payments_frame = ttk.LabelFrame(main_frame, text="Список платежей")
        payments_frame.grid(row=current_row, column=0, columnspan=2, sticky='we', padx=5, pady=10)

        payments_frame.columnconfigure(0, weight=1)
        payments_frame.rowconfigure(0, weight=1)

        cols = ('id', 'date', 'amount', 'type')
        self.tree_payments = ttk.Treeview(payments_frame, columns=cols, show='headings', height=5)

        # Настраиваем заголовки
        self.tree_payments.heading('id', text='ИД')
        self.tree_payments.heading('date', text='Дата')
        self.tree_payments.heading('amount', text='Сумма')
        self.tree_payments.heading('type', text='Тип')

        # Настраиваем ширину колонок
        self.tree_payments.column('id', width=50, anchor='center')
        self.tree_payments.column('date', width=50, anchor='center')
        self.tree_payments.column('amount', width=100, anchor='center')
        self.tree_payments.column('type', width=150, anchor='center')

        scrollbar = ttk.Scrollbar(payments_frame, orient="vertical", command=self.tree_payments.yview)
        self.tree_payments.configure(yscrollcommand=scrollbar.set)

        self.tree_payments.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')

        current_row += 1

        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.grid(row=current_row, column=0, columnspan=2, sticky='e',
                           pady=15)

        self.btn_gen_contract = ttk.Button(buttons_frame, text="Сформировать договор", command=self.generate_contract)
        self.btn_gen_contract.pack(side='left', padx=5)

        self.btn_gen_app = ttk.Button(buttons_frame, text="Сформировать заявление", command=self.generate_statement)
        self.btn_gen_app.pack(side='left', padx=5)

        self.btn_ok = ttk.Button(buttons_frame, text="OK", command=self.on_ok)
        self.btn_ok.pack(side='right', padx=5)

        self.btn_cancel = ttk.Button(buttons_frame, text="Отмена", command=self.on_cancel)
        self.btn_cancel.pack(side='right', padx=5)

        self.tree_payments.bind("<Double-Button-1>", self._get_receipt_client)

        self.transient(parent)

    def set_data_client(self, client: ClientCard):
        """Функция заполняет данные абонента из базы в Карточку абонента.
        :param client: Базовая модель Клиента.
        """
        passport_client = client.passport

        self.personal_account_entry.insert(0, int(client.personal_account))
        self.full_name_entry.insert(0, client.full_name)
        self.text_address.insert(0, client.address)
        self.phone_entry.insert(0, client.phone_number)
        self.tariff_entry.current(self._get_tariffs().index(client.tariff))
        self.balance_entry.insert(0, float(client.balance))
        self.combo_status.current(self.status_list.index(client.status))
        if client.status_date:
            self.status_date_entry.set_date(client.status_date.date())
        self.connect_date_entry.set_date(client.connection_date.date())
        self.passport_ser_num.insert(0, passport_client.get("ser_num", "Нет"))
        self.passport_data.insert(0, passport_client.get("date", "Нет"))
        self.passport_how.insert(0, passport_client.get("how", "Нет"))

        for item in self.tree_accruals.get_children():
            self.tree_accruals.delete(item)

        for db in get_db():
            for accrual in get_accruals_by_client(db, client_id=client.client_id):
                (self.tree_accruals.insert("", "end", values=(
                    accrual.created_at.strftime("%d.%m.%Y"),
                    accrual.amount,
                    accrual.accrual_date.month,
                )))
            break

        for item in self.tree_payments.get_children():
            self.tree_payments.delete(item)

        for db in get_db():
            for payment in get_payments_by_client(db, client_id=client.client_id):
                (self.tree_payments.insert("", "end", values=(
                    payment.id,
                    payment.payment_date.strftime("%d.%m.%Y"),
                    payment.amount,
                    payment.status.title(),
                )))
            break

    def on_ok(self):
        """Функция сохранения данных из карточки Абонента."""
        # Получаем данные из формы Карточка Абонента
        personal_account_client = self.personal_account_entry.get()  # Получаем лицевой счет Абонента

        passport = {
            "ser_num": self.passport_ser_num.get(),
            "date": self.passport_data.get(),
            "how": self.passport_how.get(),
        }
        current_client = ClientUpdate(
            full_name=self.full_name_entry.get(),
            address=self.text_address.get(),
            phone_number=self.phone_entry.get(),
            tariff=self.tariff_entry.get(),
            # balance=float(self.balance_entry.get()),
            passport=passport,
            status=self.combo_status.get(),
            connection_date=self.connect_date_entry.get_date(),
            status_date=self.status_date_entry.get_date(),
        )
        try:
            for db in get_db():
                client = get_client_by_pa(db, int(personal_account_client))
                update_client(db, client.id, current_client)
                break
        except Exception as e:
            messagebox.showerror(
                "Ошибка!",
                f"Не удалось обновить клиента. \nПодробности: {e}"
            )
        self.destroy()  # Закрываем окно

    def on_cancel(self):
        """Функция закрытия окна"""
        self.destroy()

    def generate_contract(self):
        """Формирование договора с клиентом и сохранение готового файла."""

        full_name = self.full_name_entry.get().split()
        if not len(full_name) == 3:
            messagebox.showwarning("Внимание", "Заполните ФИО по примеру: Иванов Иван Иванович.")
            return
        first_name = full_name[0]
        last_name = full_name[1]
        other_name = full_name[2]

        passport_ser_num = self.passport_ser_num.get().split()
        if not len(passport_ser_num) == 2:
            messagebox.showwarning("Внимание", "Заполните серию и номер паспорта по примеру: 1234 567890.")
            return
        passport_ser = passport_ser_num[0]
        passport_number = passport_ser_num[1]

        full_address = self.text_address.get().split()
        if not len(full_address) == 2:
            messagebox.showwarning("Внимание", "Заполните адрес по примеру: Дзержинского 116/2.")
            return
        street = full_address[0]
        house = None
        apartment = None
        house_and_apartment = full_address[1].split("/")
        if len(house_and_apartment) == 1:
            house = house_and_apartment[0]
            apartment = "Нет"
        elif len(house_and_apartment) == 2:
            house = house_and_apartment[0]
            apartment = house_and_apartment[1]
        else:
            messagebox.showwarning("Внимание", "Заполните адрес по примеру: Дзержинского 116/2.")
            return

        phone_number = self.phone_entry.get()

        wb = load_workbook(filename='templates/agreement.xlsx')
        sheet = wb['1']

        # Блок 'АБОНЕНТ'
        sheet['AA24'] = first_name
        sheet['AA25'] = last_name
        sheet['AA26'] = other_name

        # Блок 'Реквизиты документа удостоверяющего личность'
        sheet['AA27'] = passport_ser
        sheet['AA28'] = passport_number
        sheet['AA29'] = self.passport_data.get()
        sheet['AA30'] = self.passport_how.get()

        # Блок 'Адрес абонента'
        sheet['AA31'] = street
        sheet['AA32'] = house
        sheet['AA33'] = apartment
        sheet['AA34'] = phone_number

        # Блок 'Тарифный план'
        sheet['AD38'] = self.tariff_entry.get()

        sheet['AG95'] = self.full_name_entry.get()

        result = self._save_report(wb, f"Договор_ЛС-{self.personal_account_entry.get()}_{first_name}.xlsx")
        if result:
            # Можно, например, автоматически открыть файл после сохранения
            import os
            os.startfile(result)

    def generate_statement(self):
        """Формирование формы заявления на подключение"""

        wb = load_workbook(filename='templates/client_statement.xlsx')
        sheet = wb['1']

        sheet['O12'] = self.full_name_entry.get()
        sheet['O13'] = self.text_address.get()
        sheet['O14'] = self.phone_entry.get()

        passport_ser_num = self.passport_ser_num.get().split()
        if not len(passport_ser_num) == 2:
            messagebox.showwarning("Внимание", "Заполните серию и номер паспорта по примеру: 1234 567890.")
            return
        passport_ser = passport_ser_num[0]
        passport_number = passport_ser_num[1]

        sheet['AH16'] = passport_ser
        sheet['AH17'] = passport_number
        sheet['AH18'] = f"{self.passport_data.get()} {self.passport_how.get()}"

        tariff_name = self.tariff_entry.get()
        tariff = None
        service = None # Услуга должна называться 'Подключение'
        db = next(get_db())
        try:
            tariff = get_tariff_by_name(db, tariff_name)
            service = get_service_by_name(db, "Подключение")
        finally:
            db.close()

        if not tariff:
            messagebox.showwarning("Внимание", "Тариф не найден.")
            return

        if not service:
            messagebox.showwarning("Ошибка", "Такой услуги нет, ее необходимо добавить. Услуга должна называться 'Подключение'")
            return

        sheet['C32'] = tariff.name
        sheet['L32'] = service.service_price
        sheet['Z32'] = tariff.monthly_price

        sheet['AG38'] = datetime.now().strftime("%d.%m.%Y")

        result = self._save_report(wb, f"Заявление_ЛС-{self.personal_account_entry.get()}_{self.full_name_entry.get()}.xlsx")
        if result:
            # Можно, например, автоматически открыть файл после сохранения
            import os
            os.startfile(result)

    def _get_tariffs(self):
        """Получает тарифы из базы и возвращает списком"""
        try:
            list_tariffs = []
            for db in get_db():
                tariffs = get_tariffs(db)
                if tariffs:

                    for tariff in tariffs:
                        list_tariffs.append(tariff.name)
                    return list_tariffs
                else:
                    list_tariffs[0] = "Нет"
                break
            return list_tariffs
        except Exception as e:
            messagebox.showerror(
                "Ошибка!",
                f"Возникла ошибка!\nПодробности: \n{e}"
            )

    def _get_receipt_client(self, event):
        """
        Получение квитанции платежа.
        :param event: Двойное нажатие мышки
        :return:
        """
        item_id = self.tree_payments.identify_row(event.y)
        if not item_id:
            return
        item_data = self.tree_payments.item(item_id)
        values = item_data['values']
        payment = None
        payment_id = int(values[0])
        if payment_id:
            for db in get_db():
                payment = get_payment_by_id(db, payment_id)
                break
        wb = load_workbook(filename='templates/receipt.xlsx')
        sheet = wb['1']
        payment_month = None
        personal_account = self.personal_account_entry.get()
        if payment:
            payment_month = payment.payment_date.month
            formatted_date = payment.created_at.strftime("%d.%m.%Y") if payment.created_at else ""
            sheet['J3'] = formatted_date
            sheet['W3'] = formatted_date
            full_name = self.full_name_entry.get().upper()
            sheet['B4'] = full_name
            sheet['O4'] = full_name
            address = self.text_address.get().upper()
            sheet['B5'] = address
            sheet['O5'] = address

            sheet['J6'] = personal_account
            sheet['W6'] = personal_account
            id_payment = payment.id
            sheet['J8'] = id_payment
            sheet['W8'] = id_payment
            payment_data_start, payment_data_end = self._get_month_range(payment.payment_date.month,
                                                                         payment.payment_date.year)
            sheet['C11'] = payment_data_start.strftime("%d.%m.%Y")
            sheet['I11'] = payment_data_end.strftime("%d.%m.%Y")
            sheet['P11'] = payment_data_start.strftime("%d.%m.%Y")
            sheet['V11'] = payment_data_end.strftime("%d.%m.%Y")
            amount = f"{payment.amount:.2f} руб."
            sheet['I12'] = amount
            sheet['V12'] = amount
            sheet['I14'] = amount
            sheet['V14'] = amount

        result = self._save_report(wb, f"Квитанция_ЛС-{personal_account}_ИД-{payment_id}_месяц-{payment_month}.xlsx")
        if result:
            # Можно, например, автоматически открыть файл после сохранения
            import os
            os.startfile(result)

    def _get_month_range(self, month_num: int, year: int) -> tuple:
        """
        Метод получения периода полного месяца.
        :param month_num: Номер месяца (1-12)
        :param year: Год
        :return: Кортеж объектов дат, например: data_start, data_end
        """
        if not 1 <= month_num <= 12:
            return " ", " "

        start_date = date(year, month_num, 1)

        _, last_day = calendar.monthrange(year, month_num)

        end_date = date(year, month_num, last_day)

        return start_date, end_date

    def _save_report(self, wb, default_filename="Отчет.xlsx"):
        """
        Метод сохранения файл отчетов, заявлений и договоров.
        :param wb: Объект шаблона книги
        :param default_filename: Имя файла
        :return:
        """
        file_path_str = filedialog.asksaveasfilename(
            title=f"Выберите место для сохранения {default_filename}",
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        # Если пользователь нажал "Отмена", file_path_str будет пустой строкой
        if not file_path_str:
            messagebox.showinfo(
                title="Отмена",
                message="Операция отменена"
            )
            return None

        # 2. Используем pathlib для работы с путем
        save_path = Path(file_path_str)

        try:
            # Убеждаемся, что директория существует (на случай, если путь введен вручную)
            save_path.parent.mkdir(parents=True, exist_ok=True)

            # 3. Сохраняем книгу
            wb.save(str(save_path))

            messagebox.showinfo("Успех", f"Файл сохранен:\n{save_path.name}")
            return save_path

        except PermissionError:
            error_msg = "Ошибка доступа: закройте файл, если он открыт в Excel, и попробуйте снова."
            messagebox.showerror("Ошибка доступа", error_msg)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")

        return None


class WindowReport(tkinter.Toplevel):
    def __init__(self, parent, title: str = "Отчет", report_type: int = 0, start_date=None, end_date=None):
        super().__init__(parent)
        self.title(title)
        self.geometry("800x650")
        self.resizable(True, True)

        self.report_type = report_type
        self.start_date = start_date
        self.end_date = end_date

        self.total_amount_var = tkinter.StringVar(value="0.00")

        total_frame = ttk.Frame(self, padding=10, relief="flat")
        total_frame.pack(side="bottom", fill="x")
        ttk.Button(total_frame, text="Выгрузить в Excel",
                   command=self._export_to_excel).pack(side="right", padx=10)
        if report_type == 0:
            ttk.Label(total_frame, text="Итоговая сумма задолженности:").pack(side="left")
            ttk.Label(total_frame, textvariable=self.total_amount_var,
                      foreground="blue").pack(side="left", padx=5)
            ttk.Label(total_frame, text="руб.").pack(side="left")
        elif report_type == 1:
            ttk.Label(total_frame, text="Итоговая сумма за период:").pack(side="left")
            ttk.Label(total_frame, textvariable=self.total_amount_var,
                      foreground="blue").pack(side="left", padx=5)
            ttk.Label(total_frame, text="руб.").pack(side="left")

        report_frame = ttk.Frame(self, padding=5)
        report_frame.pack(fill="both", expand=True)

        if report_type == 0:
            cols = {
                "personal_account": ("Л/С", 100),
                "full_name": ("ФИО", 200),
                "address": ("Адрес", 250),
                "balance": ("Баланс", 100),
                "status": ("Статус", 100),
            }
        else:
            cols = {
                "personal_account": ("Л/С", 100),
                "full_name": ("ФИО", 250),
                "payment_date": ("Дата платежа", 120),
                "amount": ("Сумма", 100),
            }

        self.tree_frame = ttk.Treeview(report_frame, columns=list(cols.keys()), show='headings')

        scrollbar = ttk.Scrollbar(report_frame, orient="vertical", command=self.tree_frame.yview)
        self.tree_frame.configure(yscrollcommand=scrollbar.set)

        for col_key, (col_text, col_width) in cols.items():
            self.tree_frame.heading(col_key, text=col_text)
            self.tree_frame.column(col_key, width=col_width, anchor="center")

        self.tree_frame.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        if report_type == 0:
            self._load_clients()
        elif report_type == 1:
            self._load_payments()

        self._center_to_parent(parent)

        self.transient(parent)
        self.grab_set()

    def _center_to_parent(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (self.winfo_width() // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

    def _load_clients(self):
        """
        Загружает и отображает список клиентов должников.
        """
        # 1. Очистка Treeview
        for item in self.tree_frame.get_children():
            self.tree_frame.delete(item)

        clients = None
        for db in get_db():
            clients = get_debtors_report(db)
            break

        # 3. Отображение
        self._display_clients(clients)

    def _load_payments(self):
        """
        Загружает и отображает список платежей за период.
        """
        # 1. Очистка Treeview
        for item in self.tree_frame.get_children():
            self.tree_frame.delete(item)

        payments = None
        for db in get_db():
            payments = get_payments_in_range(db, self.start_date, self.end_date)
            break

        # 3. Отображение
        self._display_payments(payments)

    def _display_clients(self, clients):
        """Отображает список должников (имеющих отрицательный баланс)."""
        # Очистка Treeview
        for item in self.tree_frame.get_children():
            self.tree_frame.delete(item)

        total_sum = 0.0

        for client in clients:
            self.tree_frame.insert("", "end", values=(
                client.personal_account,
                client.full_name,
                client.address,
                f"{client.balance:.2f}",  # Форматируем баланс
                client.status.value,
            ))
            total_sum += float(client.balance)

        self.total_amount_var.set(f"{total_sum:,.2f}".replace(",", " "))

    def _display_payments(self, payments):
        """Отображает список платежей и обновляет итог."""
        for item in self.tree_frame.get_children():
            self.tree_frame.delete(item)

        total_sum = 0.0

        for db in get_db():
            for payment in payments:
                client = get_client_by_id(db, payment.client_id)
                self.tree_frame.insert("", "end", values=(
                    client.personal_account,
                    client.full_name,
                    payment.payment_date.strftime("%d.%m.%Y"),
                    f"{payment.amount:.2f}",
                ))
                total_sum += float(payment.amount)
            break

        self.total_amount_var.set(f"{total_sum:,.2f}".replace(",", " "))

    def _export_to_excel(self):
        # 1. Сбор данных из Treeview
        data = []
        columns = [self.tree_frame.heading(col)["text"] for col in self.tree_frame["columns"]]

        for item in self.tree_frame.get_children():
            data.append(self.tree_frame.item(item)["values"])

        if not data:
            messagebox.showwarning("Внимание", "Нет данных для выгрузки")
            return

        # 2. Выбор пути сохранения
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Сохранить отчет"
        )

        if file_path:
            try:
                df = pd.DataFrame(data, columns=columns)
                # Добавляем строку с итогом в конец файла
                total_label = "Итоговая сумма:"
                total_val = self.total_amount_var.get()

                # Создаем Excel-файл
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Отчет')

                    # Небольшое форматирование (опционально)
                    ws = writer.sheets['Отчет']
                    ws.append([])  # пустая строка
                    ws.append([total_label, total_val])

                messagebox.showinfo("Успех", f"Отчет успешно сохранен в:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")


if __name__ == "__main__":
    app = BillingSysemApp()
    app.mainloop()
